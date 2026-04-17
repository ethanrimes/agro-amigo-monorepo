#!/usr/bin/env python3
"""
Upload official DANE CPC 2.1 A.C. classification data to Supabase.

Sources:
  - Goods (Sections 0-4): CPC-21AC-BienesTransportablesSec04-2023.xlsx
  - Services (Sections 5-9): CPC_21AC_Servicios_Sec_5_9_2022.xlsx

Both files are publicly available from:
  https://www.dane.gov.co/index.php/sistema-estadistico-nacional-sen/
    normas-y-estandares/nomenclaturas-y-clasificaciones/clasificaciones/
    clasificacion-central-de-productos-cpc
"""

import os
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass

import openpyxl
import requests

_parent_dir = str(Path(__file__).parent.parent)
if _parent_dir not in sys.path:
    sys.path.insert(0, _parent_dir)

from backend.supabase_client import get_db_connection

GOODS_URL = "https://www.dane.gov.co/files/sen/nomenclatura/cpc/CPC-21AC-BienesTransportablesSec04-2023.xlsx"
SERVICES_URL = "https://www.dane.gov.co/files/sen/nomenclatura/cpc/CPC_21AC_Servicios_Sec_5_9_2022.xlsx"

CPC_DIR = Path(__file__).parent.parent / "exports" / "cpc"


@dataclass
class CpcEntry:
    code: str
    title: str
    level: str  # section, division, group, class, subclass, product
    parent_code: Optional[str]
    section_code: Optional[str]
    division_code: Optional[str]
    group_code: Optional[str]
    class_code: Optional[str]


def download_file(url: str, dest: Path) -> Path:
    """Download a file if not already present."""
    if dest.exists():
        print(f"  [LOCAL] Already on disk: {dest.name}")
        return dest
    print(f"  Downloading: {dest.name} ...", end="", flush=True)
    resp = requests.get(url, timeout=120)
    resp.raise_for_status()
    dest.write_bytes(resp.content)
    size_mb = len(resp.content) / 1024 / 1024
    print(f" {size_mb:.1f} MB")
    return dest


def parse_goods_file(file_path: Path) -> List[CpcEntry]:
    """Parse CPC goods file (sections 0-4). Has Id column at position 0."""
    wb = openpyxl.load_workbook(str(file_path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    entries = []
    # Track current ancestors for hierarchy
    cur_section = cur_division = cur_group = cur_class = cur_subclass = None

    for row in ws.iter_rows(min_row=7, values_only=True):
        # Goods file: [Id, Grupo, Clase, Subclase, Título, Ud, CIIU, SA, CPC2AC, ...]
        grupo_col = str(row[1] or "").strip()
        clase_col = str(row[2] or "").strip()
        subclase_col = str(row[3] or "").strip()
        titulo = str(row[4] or "").strip()

        if not titulo:
            continue

        entry = _classify_row(
            grupo_col, clase_col, subclase_col, titulo,
            cur_section, cur_division, cur_group, cur_class, cur_subclass,
        )
        if entry:
            entries.append(entry)
            # Update ancestor tracking
            if entry.level == "section":
                cur_section = entry.code
                cur_division = cur_group = cur_class = cur_subclass = None
            elif entry.level == "division":
                cur_division = entry.code
                cur_group = cur_class = cur_subclass = None
            elif entry.level == "group":
                cur_group = entry.code
                cur_class = cur_subclass = None
            elif entry.level == "class":
                cur_class = entry.code
                cur_subclass = None
            elif entry.level == "subclass":
                cur_subclass = entry.code

    wb.close()
    return entries


def parse_services_file(file_path: Path) -> List[CpcEntry]:
    """Parse CPC services file (sections 5-9). No Id column."""
    wb = openpyxl.load_workbook(str(file_path), read_only=True)
    ws = wb[wb.sheetnames[0]]
    entries = []
    cur_section = cur_division = cur_group = cur_class = cur_subclass = None

    for row in ws.iter_rows(min_row=7, values_only=True):
        # Services file: [Grupo, Clase, Subclase, Título, CIIU, CPC2AC]
        grupo_col = str(row[0] or "").strip()
        clase_col = str(row[1] or "").strip()
        subclase_col = str(row[2] or "").strip()
        titulo = str(row[3] or "").strip()

        if not titulo:
            continue

        entry = _classify_row(
            grupo_col, clase_col, subclase_col, titulo,
            cur_section, cur_division, cur_group, cur_class, cur_subclass,
        )
        if entry:
            entries.append(entry)
            if entry.level == "section":
                cur_section = entry.code
                cur_division = cur_group = cur_class = cur_subclass = None
            elif entry.level == "division":
                cur_division = entry.code
                cur_group = cur_class = cur_subclass = None
            elif entry.level == "group":
                cur_group = entry.code
                cur_class = cur_subclass = None
            elif entry.level == "class":
                cur_class = entry.code
                cur_subclass = None
            elif entry.level == "subclass":
                cur_subclass = entry.code

    wb.close()
    return entries


def _classify_row(
    grupo_col: str, clase_col: str, subclase_col: str, titulo: str,
    cur_section, cur_division, cur_group, cur_class, cur_subclass,
) -> Optional[CpcEntry]:
    """Classify a row into the CPC hierarchy."""

    # SECCIÓN (e.g., "SECCIÓN 0")
    m = re.match(r"SECCI[ÓO]N\s+(\d+)", grupo_col, re.IGNORECASE)
    if m:
        code = m.group(1)
        return CpcEntry(
            code=code, title=titulo, level="section",
            parent_code=None,
            section_code=code, division_code=None, group_code=None, class_code=None,
        )

    # DIVISIÓN (e.g., "DIVISIÓN 01")
    m = re.match(r"DIVISI[ÓO]N\s+(\d+)", grupo_col, re.IGNORECASE)
    if m:
        code = m.group(1)
        return CpcEntry(
            code=code, title=titulo, level="division",
            parent_code=cur_section,
            section_code=cur_section, division_code=code, group_code=None, class_code=None,
        )

    # Group (3-digit code in grupo_col, nothing in clase/subclase)
    if re.match(r"^\d{3}$", grupo_col) and not clase_col and not subclase_col:
        return CpcEntry(
            code=grupo_col, title=titulo, level="group",
            parent_code=cur_division,
            section_code=cur_section, division_code=cur_division,
            group_code=grupo_col, class_code=None,
        )

    # Class (4-digit code in clase_col, nothing in subclase)
    if clase_col and re.match(r"^\d{4}$", clase_col) and not subclase_col:
        return CpcEntry(
            code=clase_col, title=titulo, level="class",
            parent_code=cur_group,
            section_code=cur_section, division_code=cur_division,
            group_code=cur_group, class_code=clase_col,
        )

    # Subclass or Product (in subclase_col)
    if subclase_col and re.match(r"^\d+$", subclase_col):
        code = subclase_col
        if len(code) <= 5:
            # 5-digit subclass
            return CpcEntry(
                code=code, title=titulo, level="subclass",
                parent_code=cur_class,
                section_code=cur_section, division_code=cur_division,
                group_code=cur_group, class_code=cur_class,
            )
        else:
            # 7-digit product code (Colombia-specific adaptation)
            return CpcEntry(
                code=code, title=titulo, level="product",
                parent_code=cur_subclass,
                section_code=cur_section, division_code=cur_division,
                group_code=cur_group, class_code=cur_class,
            )

    return None


def upload_to_supabase(entries: List[CpcEntry]):
    """Insert CPC entries into dim_cpc table."""
    conn = get_db_connection()
    cur = conn.cursor()

    # We need to insert in order (parents before children) since we have FK constraints
    # The entries are already in document order which respects the hierarchy
    inserted = 0
    skipped = 0

    for entry in entries:
        try:
            cur.execute("""
                INSERT INTO dim_cpc (code, title, level, parent_code, section_code, division_code, group_code, class_code)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (code) DO UPDATE SET
                    title = EXCLUDED.title,
                    level = EXCLUDED.level,
                    parent_code = EXCLUDED.parent_code,
                    section_code = EXCLUDED.section_code,
                    division_code = EXCLUDED.division_code,
                    group_code = EXCLUDED.group_code,
                    class_code = EXCLUDED.class_code
            """, (
                entry.code, entry.title, entry.level, entry.parent_code,
                entry.section_code, entry.division_code, entry.group_code, entry.class_code,
            ))
            inserted += 1
        except Exception as e:
            print(f"  [WARN] Failed to insert {entry.code}: {e}")
            conn.rollback()
            skipped += 1
            continue

    conn.commit()
    print(f"  Inserted/updated: {inserted}, Skipped: {skipped}")
    return inserted


def link_insumos_to_cpc(conn):
    """
    Set dim_insumo.cpc_id from existing cpc_code values.
    Sources: dim_insumo.cpc_code, insumo_prices_department.cpc_code
    """
    cur = conn.cursor()

    # First, update from dim_insumo.cpc_code where it already has a value
    cur.execute("""
        UPDATE dim_insumo
        SET cpc_id = cpc_code
        WHERE cpc_code IS NOT NULL
          AND cpc_code != ''
          AND cpc_id IS NULL
          AND EXISTS (SELECT 1 FROM dim_cpc WHERE code = dim_insumo.cpc_code)
    """)
    direct = cur.rowcount
    print(f"  Linked {direct} insumos directly from dim_insumo.cpc_code")

    # Second, for insumos without cpc_code, try to get it from their dept price rows
    cur.execute("""
        UPDATE dim_insumo di
        SET cpc_id = sub.cpc_code
        FROM (
            SELECT DISTINCT ON (insumo_id) insumo_id, cpc_code
            FROM insumo_prices_department
            WHERE cpc_code IS NOT NULL AND cpc_code != ''
            ORDER BY insumo_id, price_date DESC
        ) sub
        WHERE di.id = sub.insumo_id
          AND di.cpc_id IS NULL
          AND EXISTS (SELECT 1 FROM dim_cpc WHERE code = sub.cpc_code)
    """)
    from_prices = cur.rowcount
    print(f"  Linked {from_prices} insumos from department price CPC codes")

    # Also backfill dim_insumo.cpc_code where it was NULL but we found it
    cur.execute("""
        UPDATE dim_insumo
        SET cpc_code = cpc_id
        WHERE cpc_id IS NOT NULL AND (cpc_code IS NULL OR cpc_code = '')
    """)
    backfilled = cur.rowcount
    print(f"  Backfilled {backfilled} dim_insumo.cpc_code values")

    conn.commit()

    # Report stats
    cur.execute("SELECT COUNT(*) as total FROM dim_insumo")
    total = cur.fetchone()["total"]
    cur.execute("SELECT COUNT(*) as linked FROM dim_insumo WHERE cpc_id IS NOT NULL")
    linked = cur.fetchone()["linked"]
    print(f"  Total insumos: {total}, Linked to CPC: {linked} ({100*linked/total:.1f}%)")


def main():
    print("=" * 60)
    print("CPC 2.1 A.C. Classification Uploader")
    print("=" * 60)

    # Ensure directory exists
    CPC_DIR.mkdir(parents=True, exist_ok=True)

    # Download files
    print("\n1. Downloading CPC classification files...")
    goods_path = download_file(GOODS_URL, CPC_DIR / "CPC-21AC-BienesTransportablesSec04-2023.xlsx")
    services_path = download_file(SERVICES_URL, CPC_DIR / "CPC_21AC_Servicios_Sec_5_9_2022.xlsx")

    # Parse files
    print("\n2. Parsing goods classification (Sections 0-4)...")
    goods_entries = parse_goods_file(goods_path)
    print(f"  Parsed {len(goods_entries)} entries")

    print("\n3. Parsing services classification (Sections 5-9)...")
    services_entries = parse_services_file(services_path)
    print(f"  Parsed {len(services_entries)} entries")

    all_entries = goods_entries + services_entries

    # Count by level
    by_level = {}
    for e in all_entries:
        by_level[e.level] = by_level.get(e.level, 0) + 1
    print(f"\n  Total entries: {len(all_entries)}")
    for level in ["section", "division", "group", "class", "subclass", "product"]:
        print(f"    {level}: {by_level.get(level, 0)}")

    # Run migration first
    print("\n4. Running migration (creating dim_cpc table)...")
    conn = get_db_connection()
    migration_path = Path(__file__).parent / "024_create_cpc_table.sql"
    sql = migration_path.read_text(encoding="utf-8")
    cur = conn.cursor()
    cur.execute(sql)
    conn.commit()
    print("  Migration applied.")

    # Upload
    print("\n5. Uploading CPC data to Supabase...")
    upload_to_supabase(all_entries)

    # Link insumos
    print("\n6. Linking insumos to CPC codes...")
    link_insumos_to_cpc(conn)

    print("\n" + "=" * 60)
    print("Done!")


if __name__ == "__main__":
    main()
