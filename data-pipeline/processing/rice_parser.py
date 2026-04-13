#!/usr/bin/env python3
"""
Parser for SIPSA rice mill price Excel files.

Two formats:
- Old (2013-2022): Año, Mes, Producto, Código Municipio, Municipio, Codigo Depto, Departamento, Precio
- New (2023+): Fecha (datetime), Producto, Código Municipio, Municipio, Codigo Depto, Departamento, Precio
"""

import sys
from datetime import datetime, date
from typing import List, Tuple, Optional
from pathlib import Path

import openpyxl

_parent_dir = str(Path(__file__).parent.parent)
if _parent_dir not in sys.path:
    sys.path.insert(0, _parent_dir)

from config import MONTH_ABBR_MAP
from backend.database import ProcessedPrice, ProcessingError

RICE_CATEGORY = 'Granos y cereales'
RICE_SUBCATEGORY = 'Arroz en molino'
RICE_PRESENTATION = 'Tonelada'
RICE_UNITS = '1 Tonelada'


class RiceParser:
    """Parser for SIPSA rice mill price Excel files."""

    def __init__(self, download_entry_id: Optional[str] = None):
        self.download_entry_id = download_entry_id

    def parse(self, file_path: str, storage_path: str) -> Tuple[List[ProcessedPrice], List[ProcessingError]]:
        try:
            wb = openpyxl.load_workbook(file_path, read_only=True)
            ws = wb.active

            fmt = self._detect_format(ws)
            print(f"    Rice format: {fmt}")

            if fmt == 'new':
                prices = self._parse_new(ws, storage_path)
            else:
                prices = self._parse_old(ws, storage_path)

            wb.close()
            return prices, []

        except Exception as e:
            return [], [ProcessingError(
                error_type='excel_parse_error',
                error_message=str(e),
                source_path=storage_path,
                source_type='excel',
                download_entry_id=self.download_entry_id
            )]

    def _detect_format(self, ws) -> str:
        """Old format has 'Año' in column A header; new has 'Fecha'."""
        for row in ws.iter_rows(min_row=6, max_row=8, values_only=True):
            for val in row:
                if val and 'fecha' in str(val).lower():
                    return 'new'
                if val and str(val).lower().strip() == 'año':
                    return 'old'
        return 'new'

    def _parse_new(self, ws, storage_path: str) -> List[ProcessedPrice]:
        """New format: Fecha, Producto, Código Muni, Municipio, Código Depto, Departamento, Precio"""
        prices = []
        for row in ws.iter_rows(min_row=8, values_only=True):
            if not row or row[0] is None or row[1] is None:
                continue
            fecha = row[0]
            producto = row[1]
            muni_name = row[3]
            precio = row[6]

            if not isinstance(fecha, (datetime, date)) or precio is None:
                continue
            try:
                price_date = fecha.date() if isinstance(fecha, datetime) else fecha
                avg = float(precio)
            except (ValueError, TypeError):
                continue

            prices.append(self._make_price(price_date, str(producto).strip(), str(muni_name or '').strip(), avg, storage_path))
        return prices

    def _parse_old(self, ws, storage_path: str) -> List[ProcessedPrice]:
        """Old format: Año, Mes, Producto, Código Muni, Municipio, Código Depto, Departamento, Precio"""
        prices = []
        for row in ws.iter_rows(min_row=8, values_only=True):
            if not row or row[0] is None or row[2] is None:
                continue
            year_val = row[0]
            month_val = row[1]
            producto = row[2]
            muni_name = row[4]
            precio = row[7]

            if precio is None:
                continue
            try:
                year = int(year_val)
                month_str = str(month_val).strip().lower()[:3]
                month = MONTH_ABBR_MAP.get(month_str)
                if not month:
                    continue
                price_date = date(year, month, 1)
                avg = float(precio)
            except (ValueError, TypeError):
                continue

            prices.append(self._make_price(price_date, str(producto).strip(), str(muni_name or '').strip(), avg, storage_path))
        return prices

    def _make_price(self, price_date, product, city, avg_price, storage_path) -> ProcessedPrice:
        return ProcessedPrice(
            category=RICE_CATEGORY,
            subcategory=RICE_SUBCATEGORY,
            product=product,
            presentation=RICE_PRESENTATION,
            units=RICE_UNITS,
            price_date=price_date,
            round=1,
            min_price=None,
            max_price=None,
            avg_price=avg_price,
            source_type='excel',
            source_path=storage_path,
            download_entry_id=self.download_entry_id,
            city=city,
            market=''
        )
