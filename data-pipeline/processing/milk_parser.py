#!/usr/bin/env python3
"""
Parser for SIPSA milk price Excel files.

Handles three different formats:
1. Historical 2013-2019: Year, Month, Department, Municipality, Avg Price
2. Historical 2020-2025: Date, Department, Municipality, Avg Price
3. Monthly 2023+: Department, Municipality, Min/Max/Avg Price
"""

import os
import re
import sys
from datetime import datetime, date
from typing import List, Optional, Tuple
from pathlib import Path

import openpyxl

_parent_dir = str(Path(__file__).parent.parent)
if _parent_dir not in sys.path:
    sys.path.insert(0, _parent_dir)

from config import MONTHS_ES_REVERSE, MONTH_ABBR_MAP
from backend.database import ProcessedPrice, ProcessingError

# Milk product constants
MILK_CATEGORY = 'Lácteos y huevos'
MILK_SUBCATEGORY = 'Leche cruda en finca'
MILK_PRODUCT = 'Leche cruda'
MILK_PRESENTATION = 'Litro'
MILK_UNITS = '1 Litro'


class MilkParser:
    """Parser for SIPSA milk price Excel files."""

    def __init__(self, download_entry_id: Optional[str] = None):
        self.download_entry_id = download_entry_id

    def parse(self, file_path: str, storage_path: str) -> Tuple[List[ProcessedPrice], List[ProcessingError]]:
        """
        Parse a milk price Excel file.

        Auto-detects the format based on column headers.

        Returns:
            Tuple of (prices, errors)
        """
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active

            # Detect format by scanning headers
            fmt = self._detect_format(ws)
            print(f"    Milk format detected: {fmt}")

            if fmt == 'historical_old':
                return self._parse_historical_old(ws, storage_path)
            elif fmt == 'historical_new':
                return self._parse_historical_new(ws, storage_path)
            elif fmt == 'monthly':
                return self._parse_monthly(ws, storage_path)
            else:
                return [], [ProcessingError(
                    error_type='excel_parse_error',
                    error_message=f'Unknown milk Excel format in {storage_path}',
                    source_path=storage_path,
                    source_type='excel',
                    download_entry_id=self.download_entry_id
                )]

        except Exception as e:
            return [], [ProcessingError(
                error_type='excel_parse_error',
                error_message=str(e),
                source_path=storage_path,
                source_type='excel',
                download_entry_id=self.download_entry_id
            )]

    def _detect_format(self, ws) -> str:
        """Detect which format this Excel file uses."""
        # Scan first 15 rows for header clues
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=False):
            for cell in row:
                val = str(cell.value or '').lower().strip()
                if 'mes y año' in val or 'mes y ano' in val:
                    return 'historical_new'
                if val == 'año' or val == 'ano':
                    # Check if next cell is 'Mes'
                    next_col = cell.column + 1
                    next_val = str(ws.cell(row=cell.row, column=next_col).value or '').lower().strip()
                    if next_val == 'mes':
                        return 'historical_old'
                if 'precio mínimo' in val or 'precio minimo' in val:
                    return 'monthly'
        return 'unknown'

    def _parse_historical_old(self, ws, storage_path: str) -> Tuple[List[ProcessedPrice], List[ProcessingError]]:
        """
        Parse 2013-2019 format:
        Columns: Año, Mes, Nombre departamento, Código departamento,
                 Nombre municipio, Código municipio, Precio promedio por litro
        Data starts at row 11.
        """
        prices = []
        errors = []

        for row in ws.iter_rows(min_row=11, values_only=False):
            year_val = row[0].value
            month_val = row[1].value
            dept_name = row[2].value
            dept_code = row[3].value
            muni_name = row[4].value
            muni_code = row[5].value
            avg_price = row[6].value

            if year_val is None or muni_name is None or avg_price is None:
                continue

            try:
                year = int(year_val)
                month = int(month_val)
                price_date = date(year, month, 1)
                avg = float(avg_price)
            except (ValueError, TypeError):
                continue

            city = str(muni_name).strip().title()
            prices.append(self._make_price(
                price_date=price_date,
                city=city,
                avg_price=avg,
                storage_path=storage_path
            ))

        return prices, errors

    def _parse_historical_new(self, ws, storage_path: str) -> Tuple[List[ProcessedPrice], List[ProcessingError]]:
        """
        Parse 2020-2025 format:
        Columns: Mes y año, Nombre departamento, Código departamento,
                 Nombre municipio, Código municipio, Precio promedio por litro
        Data starts at row 10.
        """
        prices = []
        errors = []

        for row in ws.iter_rows(min_row=10, values_only=False):
            date_val = row[0].value
            dept_name = row[1].value
            dept_code = row[2].value
            muni_name = row[3].value
            muni_code = row[4].value
            avg_price = row[5].value

            if date_val is None or muni_name is None or avg_price is None:
                continue

            # Date can be a datetime or a string
            if isinstance(date_val, datetime):
                price_date = date_val.date()
            elif isinstance(date_val, date):
                price_date = date_val
            else:
                continue

            try:
                avg = float(avg_price)
            except (ValueError, TypeError):
                continue

            city = str(muni_name).strip()
            prices.append(self._make_price(
                price_date=price_date,
                city=city,
                avg_price=avg,
                storage_path=storage_path
            ))

        return prices, errors

    def _parse_monthly(self, ws, storage_path: str) -> Tuple[List[ProcessedPrice], List[ProcessingError]]:
        """
        Parse monthly format (2023+):
        Columns: Código departamento, Nombre departamento, Código municipio,
                 Nombre municipio, Precio mínimo, Precio máximo, Precio medio, Tendencia
        Data starts at row 11.
        Date extracted from title in merged cells.
        """
        prices = []
        errors = []

        # Extract date from title (e.g., "Enero de 2026")
        price_date = self._extract_date_from_title(ws)
        if not price_date:
            errors.append(ProcessingError(
                error_type='missing_date',
                error_message='Could not extract date from milk monthly Excel title',
                source_path=storage_path,
                source_type='excel',
                download_entry_id=self.download_entry_id
            ))
            return prices, errors

        for row in ws.iter_rows(min_row=11, values_only=False):
            dept_code = row[0].value
            dept_name = row[1].value
            muni_code = row[2].value
            muni_name = row[3].value
            min_price = row[4].value
            max_price = row[5].value
            avg_price = row[6].value

            if muni_name is None or avg_price is None:
                continue

            try:
                avg = float(avg_price)
                lo = float(min_price) if min_price is not None else None
                hi = float(max_price) if max_price is not None else None
            except (ValueError, TypeError):
                continue

            city = str(muni_name).strip()
            prices.append(self._make_price(
                price_date=price_date,
                city=city,
                min_price=lo,
                max_price=hi,
                avg_price=avg,
                storage_path=storage_path
            ))

        return prices, errors

    def _extract_date_from_title(self, ws) -> Optional[date]:
        """Extract date from title cells like 'Precio de Leche Cruda en Finca\nEnero de 2026'."""
        for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
            for val in row:
                if not val:
                    continue
                text = str(val).lower()
                # Pattern: "month de year"
                match = re.search(r'([a-záéíóú]+)\s+de\s+(\d{4})', text)
                if match:
                    month_name, year = match.groups()
                    month = MONTHS_ES_REVERSE.get(month_name)
                    if month:
                        return date(int(year), month, 1)
        return None

    def _make_price(self, price_date, city, avg_price, storage_path,
                    min_price=None, max_price=None) -> ProcessedPrice:
        """Create a ProcessedPrice for milk."""
        return ProcessedPrice(
            category=MILK_CATEGORY,
            subcategory=MILK_SUBCATEGORY,
            product=MILK_PRODUCT,
            presentation=MILK_PRESENTATION,
            units=MILK_UNITS,
            price_date=price_date,
            round=1,
            min_price=min_price,
            max_price=max_price,
            avg_price=avg_price,
            source_type='excel',
            source_path=storage_path,
            download_entry_id=self.download_entry_id,
            city=city,
            market=''
        )
