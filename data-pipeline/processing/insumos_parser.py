#!/usr/bin/env python3
"""
Parser for SIPSA insumos (agricultural input) Excel files.

Handles two types:
- Municipality-level: Año, Mes, Depto, Municipio, Producto, Presentación, Precio
- Department-level: Año, Mes, Depto, CPC, Producto, Artículo, Casa Comercial, Registro ICA, Presentación, Precio

Each file has multiple sheets organized by input subcategory (Bioinsumos, Fertilizantes, etc.).
The sheet name/title contains the subgroup name; the Index sheet contains the group mapping.
"""

import re
import sys
from datetime import date
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass
from pathlib import Path

import openpyxl

_parent_dir = str(Path(__file__).parent.parent)
if _parent_dir not in sys.path:
    sys.path.insert(0, _parent_dir)

from config import MONTHS_ES_REVERSE

# Map of Spanish month names to numbers
MONTH_NAME_MAP = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
    'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
}


@dataclass
class InsumoMunRow:
    """Municipality-level insumo price row."""
    price_date: date
    dept_code: str
    dept_name: str
    muni_code: str
    muni_name: str
    product_name: str
    presentation: str
    avg_price: float
    grupo: str      # e.g., "Insumos agrícolas"
    subgrupo: str   # e.g., "Bioinsumos"


@dataclass
class InsumoDeptRow:
    """Department-level insumo price row."""
    price_date: date
    dept_code: str
    dept_name: str
    cpc_code: str
    product_name: str
    articulo: str
    casa_comercial: str
    registro_ica: str
    presentation: str
    avg_price: float
    grupo: str
    subgrupo: str


class InsumosParser:
    """Parser for SIPSA insumos Excel files."""

    def parse_municipality(self, file_path: str) -> List[InsumoMunRow]:
        """Parse municipality-level insumos file."""
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_groups = self._get_sheet_groups(wb)

        rows = []
        for sname in wb.sheetnames:
            if not re.match(r'^\d+\.\d+$', sname):
                continue
            ws = wb[sname]
            grupo, subgrupo = sheet_groups.get(sname, ('Unknown', sname))
            sheet_rows = self._parse_mun_sheet(ws, grupo, subgrupo)
            rows.extend(sheet_rows)
            if sheet_rows:
                print(f"    Sheet {sname} ({subgrupo}): {len(sheet_rows)} rows")

        wb.close()
        print(f"    Total municipality rows: {len(rows)}")
        return rows

    def parse_department(self, file_path: str) -> List[InsumoDeptRow]:
        """Parse department-level insumos file."""
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_groups = self._get_sheet_groups(wb)

        rows = []
        for sname in wb.sheetnames:
            if not re.match(r'^\d+\.\d+$', sname):
                continue
            ws = wb[sname]
            grupo, subgrupo = sheet_groups.get(sname, ('Unknown', sname))
            sheet_rows = self._parse_dept_sheet(ws, grupo, subgrupo)
            rows.extend(sheet_rows)
            if sheet_rows:
                print(f"    Sheet {sname} ({subgrupo}): {len(sheet_rows)} rows")

        wb.close()
        print(f"    Total department rows: {len(rows)}")
        return rows

    def _get_sheet_groups(self, wb) -> Dict[str, Tuple[str, str]]:
        """
        Parse the Index sheet to get group/subgroup mapping for each data sheet.

        Returns: {sheet_name: (grupo, subgrupo)}
        """
        mapping = {}
        index_name = next((s for s in wb.sheetnames if s.lower() in ['índice', 'indice']), None)
        if not index_name:
            return mapping

        ws = wb[index_name]
        current_grupo = ''

        for row in ws.iter_rows(max_row=50, values_only=True):
            # Look for group headers and subgroup entries
            col_a = str(row[0] or '').strip() if row[0] else ''
            col_b = str(row[1] or '').strip() if len(row) > 1 and row[1] else ''
            col_c = str(row[2] or '').strip() if len(row) > 2 and row[2] else ''

            # Detect group (e.g., "1." → "Insumos agrícolas")
            if re.match(r'^\d+\.$', col_a) and col_b:
                current_grupo = col_b

            # Detect subgroup (e.g., "1.1" → "Bioinsumos")
            if re.match(r'^\d+\.\d+$', col_b) and col_c:
                mapping[col_b] = (current_grupo, col_c)

        return mapping

    def _parse_mun_sheet(self, ws, grupo: str, subgrupo: str) -> List[InsumoMunRow]:
        """
        Parse a municipality-level data sheet.
        Headers at row 9-10, data from row 10 or 11.
        Columns: Año, Mes, Código depto, Nombre depto, Código muni, Nombre muni,
                 Nombre del producto (or Tipo de arriendo), Presentación (optional), Precio promedio
        """
        rows = []

        # Detect header row and column layout
        has_presentation = True
        data_start = 10

        for i, row_vals in enumerate(ws.iter_rows(min_row=8, max_row=12, values_only=True)):
            row_strs = [str(v or '').lower().strip() for v in row_vals]
            if any('año' in s for s in row_strs):
                data_start = i + 8 + 1  # Next row after header
                # Check if presentation column exists
                has_presentation = any('presentación' in s or 'presentacion' in s for s in row_strs)
                break

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            if not row or row[0] is None:
                continue

            try:
                year = int(row[0])
                month_str = str(row[1] or '').strip().lower()
                month = MONTH_NAME_MAP.get(month_str)
                if not month:
                    continue

                dept_code = str(row[2] or '').strip()
                dept_name = str(row[3] or '').strip()
                muni_code = str(row[4] or '').strip()
                muni_name = str(row[5] or '').strip()
                product_name = str(row[6] or '').strip()

                if has_presentation and len(row) >= 9:
                    presentation = str(row[7] or '').strip()
                    price = row[8]
                else:
                    presentation = ''
                    price = row[7] if len(row) >= 8 else None

                if not product_name or price is None:
                    continue

                avg_price = float(price)
                price_date = date(year, month, 1)

                rows.append(InsumoMunRow(
                    price_date=price_date,
                    dept_code=dept_code,
                    dept_name=dept_name,
                    muni_code=muni_code,
                    muni_name=muni_name,
                    product_name=product_name,
                    presentation=presentation,
                    avg_price=avg_price,
                    grupo=grupo,
                    subgrupo=subgrupo,
                ))
            except (ValueError, TypeError, IndexError):
                continue

        return rows

    def _parse_dept_sheet(self, ws, grupo: str, subgrupo: str) -> List[InsumoDeptRow]:
        """
        Parse a department-level data sheet.
        Columns: Año, Mes, Código depto, Nombre depto, Código CPC,
                 Nombre del producto, Artículo, Casa Comercial, Registro ICA,
                 Presentación del producto, Precio promedio departamento
        """
        rows = []
        data_start = 10

        for i, row_vals in enumerate(ws.iter_rows(min_row=8, max_row=12, values_only=True)):
            row_strs = [str(v or '').lower().strip() for v in row_vals]
            if any('año' in s for s in row_strs):
                data_start = i + 8 + 1
                break

        for row in ws.iter_rows(min_row=data_start, values_only=True):
            if not row or row[0] is None:
                continue

            try:
                year = int(row[0])
                month_str = str(row[1] or '').strip().lower()
                month = MONTH_NAME_MAP.get(month_str)
                if not month:
                    continue

                dept_code = str(row[2] or '').strip()
                dept_name = str(row[3] or '').strip()
                cpc_code = str(row[4] or '').strip()
                product_name = str(row[5] or '').strip()
                articulo = str(row[6] or '').strip()
                casa_comercial = str(row[7] or '').strip()
                registro_ica = str(row[8] or '').strip() if len(row) > 8 else ''
                presentation = str(row[9] or '').strip() if len(row) > 9 else ''
                price = row[10] if len(row) > 10 else None

                if not product_name or price is None:
                    continue

                avg_price = float(price)
                price_date = date(year, month, 1)

                rows.append(InsumoDeptRow(
                    price_date=price_date,
                    dept_code=dept_code,
                    dept_name=dept_name,
                    cpc_code=cpc_code,
                    product_name=product_name,
                    articulo=articulo,
                    casa_comercial=casa_comercial,
                    registro_ica=registro_ica,
                    presentation=presentation,
                    avg_price=avg_price,
                    grupo=grupo,
                    subgrupo=subgrupo,
                ))
            except (ValueError, TypeError, IndexError):
                continue

        return rows
