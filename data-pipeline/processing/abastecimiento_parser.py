#!/usr/bin/env python3
"""
Parser for SIPSA abastecimiento (supply/quantity) Excel files.

Handles two formats:
- Old (2013-2023): Sheets 1.1, 1.2 (semesters); 9 columns, no CPC code
- New (2024+): Sheets 2.1, 2.2, 2.3 (cuatrimestres); 10 columns with CPC code

Also extracts the CPC correlation table from new-format files.
"""

import os
import re
import sys
from datetime import datetime, date
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass
from pathlib import Path

import openpyxl

_parent_dir = str(Path(__file__).parent.parent)
if _parent_dir not in sys.path:
    sys.path.insert(0, _parent_dir)


@dataclass
class SupplyRow:
    """A single parsed supply observation."""
    city_market: str          # Combined "City, Market" string
    observation_date: date
    provenance_dept_code: str
    provenance_muni_code: str
    provenance_dept_name: str
    provenance_muni_name: str
    group: str                # Category / product group
    cpc_code: str             # CPC code (empty for old format)
    alimento: str             # Product/food name
    quantity_kg: float


class AbastecimientoParser:
    """Parser for SIPSA abastecimiento Excel files."""

    def parse(self, file_path: str) -> Tuple[List[SupplyRow], Dict[str, str]]:
        """
        Parse an abastecimiento Excel file.

        Returns:
            Tuple of (supply_rows, cpc_map)
            cpc_map: {alimento_name: cpc_code} from the CPC correlation sheet
        """
        wb = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames

        # Find data sheets (could be 1.x or 2.x naming)
        data_sheets = [s for s in sheet_names
                       if re.match(r'^\d+\.\d+$', s)
                       and 'metod' not in s.lower()
                       and 'F.' not in s]

        # Extract CPC mappings if available
        cpc_map = {}
        has_cpc_sheet = any('cpc' in s.lower() for s in sheet_names)
        if has_cpc_sheet:
            cpc_sheet_name = next(s for s in sheet_names if 'cpc' in s.lower())
            cpc_map = self._parse_cpc_sheet(wb[cpc_sheet_name])

        rows = []
        if not data_sheets:
            print(f"    [WARN] No recognized data sheets in {sheet_names}")
        else:
            # Detect old vs new format by checking header of first data sheet
            # New format has 'Código CPC' in header row; old format doesn't
            first_ws = wb[data_sheets[0]]
            fmt = self._detect_data_format(first_ws)

            for sname in data_sheets:
                ws = wb[sname]
                if fmt == 'new':
                    sheet_rows = self._parse_new_format(ws)
                else:
                    sheet_rows = self._parse_old_format(ws)
                rows.extend(sheet_rows)
                print(f"    Sheet {sname}: {len(sheet_rows)} rows")

        wb.close()
        print(f"    Format: {fmt}, Total rows: {len(rows)}, CPC entries: {len(cpc_map)}")
        return rows, cpc_map

    def _detect_data_format(self, ws) -> str:
        """Detect old vs new format by checking for 'Código CPC' in header rows 8-10."""
        for row in ws.iter_rows(min_row=8, max_row=10, values_only=True):
            for val in row:
                if val and 'cpc' in str(val).lower():
                    return 'new'
        return 'old'

    def _parse_new_format(self, ws) -> List[SupplyRow]:
        """
        Parse new format (2024+):
        Row 9: headers (Ciudad,Mercado | Fecha | Divipola Depto | Divipola Muni/País |
                        Departamento | Municipio/País | Grupo | Código CPC | Alimento | Cant Kg)
        Data starts row 10.
        """
        rows = []
        for row in ws.iter_rows(min_row=10, values_only=True):
            if not row or len(row) < 10:
                continue
            city_market = row[0]
            fecha = row[1]
            dept_code = row[2]
            muni_code = row[3]
            dept_name = row[4]
            muni_name = row[5]
            group = row[6]
            cpc_code = row[7]
            alimento = row[8]
            cant_kg = row[9]

            if not city_market or not alimento or cant_kg is None:
                continue

            obs_date = self._to_date(fecha)
            if not obs_date:
                continue

            try:
                qty = float(cant_kg)
            except (ValueError, TypeError):
                continue

            rows.append(SupplyRow(
                city_market=str(city_market).strip(),
                observation_date=obs_date,
                provenance_dept_code=str(dept_code or '').strip().strip("'"),
                provenance_muni_code=str(muni_code or '').strip().strip("'"),
                provenance_dept_name=str(dept_name or '').strip(),
                provenance_muni_name=str(muni_name or '').strip(),
                group=str(group or '').strip(),
                cpc_code=str(cpc_code or '').strip().strip("'"),
                alimento=str(alimento).strip(),
                quantity_kg=qty
            ))
        return rows

    def _parse_old_format(self, ws) -> List[SupplyRow]:
        """
        Parse old format (2013-2023):
        Row 10: headers (Ciudad,Mercado | Fecha | Código Depto | Código Muni |
                         Departamento | Municipio | Grupo | Alimento | Cant Kg)
        Data starts row 11.
        """
        rows = []
        for row in ws.iter_rows(min_row=11, values_only=True):
            if not row or len(row) < 9:
                continue

            city_market = row[0]
            fecha = row[1]
            dept_code = row[2]
            muni_code = row[3]
            dept_name = row[4]
            muni_name = row[5]
            group = row[6]
            alimento = row[7]
            cant_kg = row[8]

            if not city_market or not alimento or cant_kg is None:
                continue

            obs_date = self._to_date(fecha)
            if not obs_date:
                continue

            try:
                qty = float(cant_kg)
            except (ValueError, TypeError):
                continue

            rows.append(SupplyRow(
                city_market=str(city_market).strip(),
                observation_date=obs_date,
                provenance_dept_code=str(dept_code or '').strip().strip("'"),
                provenance_muni_code=str(muni_code or '').strip().strip("'"),
                provenance_dept_name=str(dept_name or '').strip(),
                provenance_muni_name=str(muni_name or '').strip(),
                group=str(group or '').strip(),
                cpc_code='',
                alimento=str(alimento).strip(),
                quantity_kg=qty
            ))
        return rows

    def _parse_cpc_sheet(self, ws) -> Dict[str, str]:
        """Parse CPC correlation sheet. Returns {alimento: cpc_code}."""
        cpc_map = {}
        for row in ws.iter_rows(min_row=10, values_only=True):
            if not row or len(row) < 5:
                continue
            alimento = row[2]  # Column C: Descripción Alimento SIPSA_A
            cpc_code = row[3]  # Column D: Código CPC
            if alimento and cpc_code:
                cpc_map[str(alimento).strip()] = str(cpc_code).strip()
        return cpc_map

    def _to_date(self, val) -> Optional[date]:
        if isinstance(val, datetime):
            return val.date()
        if isinstance(val, date):
            return val
        return None
